from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import RelatorioGerado
from django.utils import timezone
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth, TruncDay
from clientes.models import Cliente
from pagamentos.models import Pagamento, Fatura
from equipamentos.models import Contador
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import openpyxl
from io import BytesIO

def is_admin_or_financeiro(user):
    return user.is_staff or (hasattr(user, 'perfil') and user.perfil.tipo_usuario in ['ADMIN', 'FINANCEIRO'])

@login_required
@user_passes_test(is_admin_or_financeiro)
def exportar_relatorio_pdf(request, pk):
    relatorio = get_object_or_404(RelatorioGerado, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{relatorio.id}.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 100, f"Relatório: {relatorio.titulo}")
    
    p.setFont("Helvetica", 12)
    p.drawString(100, height - 130, f"Tipo: {relatorio.get_tipo_relatorio_display()}")
    p.drawString(100, height - 150, f"Período: {relatorio.periodo_inicio} a {relatorio.periodo_fim}")
    p.drawString(100, height - 170, f"Gerado por: {relatorio.gerado_por}")
    p.drawString(100, height - 190, f"Data de Geração: {relatorio.data_geracao.strftime('%d/%m/%Y %H:%M')}")
    
    p.drawString(100, height - 220, "Observações:")
    p.drawString(100, height - 240, relatorio.observacoes)

    p.showPage()
    p.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
@user_passes_test(is_admin_or_financeiro)
def exportar_relatorio_excel(request, pk):
    relatorio = get_object_or_404(RelatorioGerado, pk=pk)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ws['A1'] = "Título"
    ws['B1'] = relatorio.titulo
    ws['A2'] = "Tipo"
    ws['B2'] = relatorio.get_tipo_relatorio_display()
    ws['A3'] = "Período"
    ws['B3'] = f"{relatorio.periodo_inicio} a {relatorio.periodo_fim}"
    ws['A4'] = "Gerado por"
    ws['B4'] = relatorio.gerado_por
    ws['A5'] = "Data de Geração"
    ws['B5'] = relatorio.data_geracao.strftime('%d/%m/%Y %H:%M')
    ws['A6'] = "Observações"
    ws['B6'] = relatorio.observacoes

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{relatorio.id}.xlsx"'
    
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin_or_financeiro)
def relatorio_list(request):
    relatorios = RelatorioGerado.objects.all().order_by('-data_geracao')
    return render(request, 'relatorios/relatorio_list.html', {'relatorios': relatorios})

@login_required
@user_passes_test(is_admin_or_financeiro)
def gerar_relatorio_view(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo_relatorio')
        inicio = request.POST.get('periodo_inicio')
        fim = request.POST.get('periodo_fim')
        
        # Simulação de geração de relatório
        relatorio = RelatorioGerado.objects.create(
            titulo=f"Relatório de {dict(RelatorioGerado.TIPO_RELATORIO_CHOICES).get(tipo)}",
            tipo_relatorio=tipo,
            periodo_inicio=inicio,
            periodo_fim=fim,
            gerado_por=request.user.username,
            observacoes=f"Gerado automaticamente em {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        )
        return redirect('relatorio_list')
    
    return render(request, 'relatorios/relatorio_form.html', {
        'tipos': RelatorioGerado.TIPO_RELATORIO_CHOICES
    })

@login_required
@user_passes_test(is_admin_or_financeiro)
def estatisticas_gerais(request):
    total_clientes = Cliente.objects.count()
    total_recebido = Pagamento.objects.aggregate(Sum('valor_pago'))['valor_pago__sum'] or 0
    total_faturado = Fatura.objects.aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    faturas_vencidas = Fatura.objects.filter(status='VENCIDO').count()
    contadores_ativos = Contador.objects.filter(status='ATIVO').count()
    
    # Consumo Mensal (últimos 6 meses)
    consumo_mensal = Fatura.objects.annotate(mes=TruncMonth('data_emissao'))\
        .values('mes')\
        .annotate(total_kwh=Sum('consumo_kwh'), total_valor=Sum('valor_total'))\
        .order_by('-mes')[:6]
    
    # Receita Mensal (Pagamentos confirmados)
    receita_mensal = Pagamento.objects.annotate(mes=TruncMonth('data_pagamento'))\
        .values('mes')\
        .annotate(total_recebido=Sum('valor_pago'))\
        .order_by('-mes')[:6]
        
    # Receita Diária (últimos 15 dias)
    receita_diaria = Pagamento.objects.annotate(dia=TruncDay('data_pagamento'))\
        .values('dia')\
        .annotate(total_recebido=Sum('valor_pago'))\
        .order_by('-dia')[:15]

    # Pagamentos por Período (Filtro)
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    pagamentos_periodo = Pagamento.objects.all()
    
    if data_inicio and data_fim:
        pagamentos_periodo = pagamentos_periodo.filter(data_pagamento__date__range=[data_inicio, data_fim])
    
    total_periodo = pagamentos_periodo.aggregate(Sum('valor_pago'))['valor_pago__sum'] or 0
    pagamentos_periodo = pagamentos_periodo.order_by('-data_pagamento')[:20]

    # Clientes Devedores (Top 10 com maior dívida)
    clientes_devedores = Fatura.objects.filter(status__in=['PENDENTE', 'VENCIDO'])\
        .values('cliente__nome', 'cliente__nif')\
        .annotate(total_divida=Sum('valor_total'), faturas_count=Count('id'))\
        .order_by('-total_divida')[:10]
    
    context = {
        'total_clientes': total_clientes,
        'total_recebido': total_recebido,
        'total_faturado': total_faturado,
        'faturas_vencidas': faturas_vencidas,
        'contadores_ativos': contadores_ativos,
        'consumo_mensal': consumo_mensal,
        'receita_mensal': receita_mensal,
        'receita_diaria': receita_diaria,
        'clientes_devedores': clientes_devedores,
        'pagamentos_periodo': pagamentos_periodo,
        'total_periodo': total_periodo,
        'filtros': {'inicio': data_inicio, 'fim': data_fim}
    }
    return render(request, 'relatorios/estatisticas.html', context)
